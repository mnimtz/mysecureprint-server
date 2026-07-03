"""
Upload-Konverter (v6.7.30)
===========================
Konvertiert verschiedene Dokument-Formate zu PDF, damit der Web-Upload
nicht nur PDFs akzeptiert sondern auch Office-Dateien, Bilder und Text.

Konverter-Chain:
  - application/pdf                → passthrough (keine Konvertierung nötig)
  - image/png, jpg, gif, bmp, tiff → Pillow → PDF
  - text/plain                     → Pillow/einfacher Renderer → PDF
  - docx, xlsx, pptx, odt, ods, odp, rtf
    + alles andere Office-ähnliche → LibreOffice headless → PDF

Alle Konverter arbeiten über ein temporäres Verzeichnis und geben die
fertige PDF als bytes zurück. Bei Fehler wird eine ConversionError
ausgelöst die der Caller als 502/Benutzer-Feedback weiterreichen kann.
"""

from __future__ import annotations

import logging
import os
import shutil
import subprocess
import tempfile
from typing import Optional

logger = logging.getLogger("printix.upload_converter")


class ConversionError(Exception):
    """Konvertierung fehlgeschlagen — Nachricht user-lesbar."""


# ─── Format-Detection ────────────────────────────────────────────────────────

# Magic-Bytes → (mime, file_extension-hint)
_MAGIC_SIGNATURES: list[tuple[bytes, str, str]] = [
    (b"%PDF",                         "application/pdf",            "pdf"),
    (b"\x89PNG\r\n\x1a\n",            "image/png",                  "png"),
    (b"\xff\xd8\xff",                 "image/jpeg",                 "jpg"),
    (b"GIF87a",                       "image/gif",                  "gif"),
    (b"GIF89a",                       "image/gif",                  "gif"),
    (b"BM",                           "image/bmp",                  "bmp"),
    (b"II*\x00",                      "image/tiff",                 "tif"),
    (b"MM\x00*",                      "image/tiff",                 "tif"),
    (b"PK\x03\x04",                   "application/zip",            "zip"),  # docx/xlsx/pptx/odt start so
    (b"{\\rtf",                       "application/rtf",            "rtf"),
]


def detect_format(data: bytes, filename_hint: str = "") -> tuple[str, str]:
    """Ermittelt (mime, extension) anhand Magic-Bytes + Dateinamen.

    Für ZIP-basierte Office-Dokumente unterscheiden wir via Dateinamens-
    Endung zwischen docx/xlsx/pptx/odt/etc., weil der ZIP-Header gleich ist.
    """
    if len(data) < 4:
        return "application/octet-stream", ""
    for magic, mime, ext in _MAGIC_SIGNATURES:
        if data.startswith(magic):
            if mime == "application/zip":
                # Office-Format aus Dateinamen ableiten
                name = (filename_hint or "").lower()
                office_map = {
                    ".docx": ("application/vnd.openxmlformats-officedocument.wordprocessingml.document", "docx"),
                    ".xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",       "xlsx"),
                    ".pptx": ("application/vnd.openxmlformats-officedocument.presentationml.presentation","pptx"),
                    ".odt":  ("application/vnd.oasis.opendocument.text",                                 "odt"),
                    ".ods":  ("application/vnd.oasis.opendocument.spreadsheet",                          "ods"),
                    ".odp":  ("application/vnd.oasis.opendocument.presentation",                        "odp"),
                }
                for sfx, (m, e) in office_map.items():
                    if name.endswith(sfx):
                        return m, e
                return mime, ext
            return mime, ext
    # Fallback: Plaintext? Alles ASCII/UTF-8-druckbar → text
    try:
        sample = data[:4096].decode("utf-8")
        if all(c.isprintable() or c in "\r\n\t" for c in sample):
            return "text/plain", "txt"
    except UnicodeDecodeError:
        pass
    return "application/octet-stream", ""


# ─── Konverter-Implementierungen ─────────────────────────────────────────────

def _convert_image_to_pdf(data: bytes, bw: bool = False, image_size: str = "full") -> bytes:
    """PNG/JPG/GIF/BMP/TIFF → PDF via Pillow, auf A4 platziert.

    bw=True:         Bild in Graustufen konvertieren.
    image_size:
      "full"     Bild proportional auf gesamte A4-Seite skalieren (Default).
      "10x13"    Fotodruck-Format 10×13 cm, zentriert auf A4.
      "13x18"    Fotodruck-Format 13×18 cm, zentriert auf A4.
      "original" Originalgröße beibehalten (nur verkleinern wenn größer als A4).
    """
    try:
        from PIL import Image
    except ImportError as e:
        raise ConversionError("Pillow (python3-pil) nicht installiert") from e
    import io

    # A4 bei 150 dpi: 1240 × 1754 px (Portrait)
    A4_W, A4_H = 1240, 1754
    MARGIN = 40  # px Rand für "full" und "original"
    PX_PER_CM = 150 / 2.54  # ≈ 59.06 px/cm

    img = Image.open(io.BytesIO(data))
    if img.mode in ("RGBA", "LA"):
        bg = Image.new("RGB", img.size, "white")
        bg.paste(img, mask=img.split()[-1])
        img = bg
    elif img.mode == "P":
        img = img.convert("RGB")
    elif img.mode != "RGB":
        img = img.convert("RGB")
    if bw:
        img = img.convert("L").convert("RGB")

    size = image_size.strip().lower()
    if size == "10x13":
        box_w = int(10 * PX_PER_CM)  # 591 px
        box_h = int(13 * PX_PER_CM)  # 768 px
        img.thumbnail((box_w, box_h), Image.LANCZOS)
    elif size == "13x18":
        box_w = int(13 * PX_PER_CM)  # 768 px
        box_h = int(18 * PX_PER_CM)  # 1063 px
        img.thumbnail((box_w, box_h), Image.LANCZOS)
    elif size == "original":
        # Nur verkleinern wenn größer als A4 minus Rand; niemals hochskalieren
        max_w = A4_W - 2 * MARGIN
        max_h = A4_H - 2 * MARGIN
        if img.width > max_w or img.height > max_h:
            img.thumbnail((max_w, max_h), Image.LANCZOS)
    else:
        # "full": Bild proportional auf gesamte nutzbare Fläche skalieren
        img.thumbnail((A4_W - 2 * MARGIN, A4_H - 2 * MARGIN), Image.LANCZOS)

    page = Image.new("RGB", (A4_W, A4_H), "white")
    page.paste(img, ((A4_W - img.width) // 2, (A4_H - img.height) // 2))

    out = io.BytesIO()
    page.save(out, format="PDF", resolution=150.0)
    return out.getvalue()


def _convert_text_to_pdf(data: bytes) -> bytes:
    """Plaintext → mehrseitiges PDF via Pillow (Monospaced-Layout).

    Erstellt so viele A4-Seiten wie nötig — vorher wurde der Text nach
    einer Seite einfach abgeschnitten.
    """
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError as e:
        raise ConversionError("Pillow (python3-pil) nicht installiert") from e
    import io
    text = data.decode("utf-8", errors="replace")
    W, H = 1240, 1754  # A4 @150dpi
    margin = 60
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", 14)
        line_height = 20
        max_chars = (W - 2 * margin) // 8
    except Exception:
        font = ImageFont.load_default()
        line_height = 18
        max_chars = (W - 2 * margin) // 7

    # Alle zu rendernden Zeilen aufbauen (mit Umbruch)
    all_lines: list[str] = []
    for raw in text.splitlines() or [""]:
        if not raw:
            all_lines.append("")
            continue
        while raw:
            all_lines.append(raw[:max_chars])
            raw = raw[max_chars:]

    # Auf mehrere Seiten verteilen
    lines_per_page = (H - 2 * margin) // line_height
    pages_data: list[bytes] = []
    for page_start in range(0, max(1, len(all_lines)), lines_per_page):
        page_lines = all_lines[page_start:page_start + lines_per_page]
        img = Image.new("RGB", (W, H), "white")
        draw = ImageDraw.Draw(img)
        for i, line in enumerate(page_lines):
            draw.text((margin, margin + i * line_height), line, fill="black", font=font)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        pages_data.append(buf.getvalue())

    # Alle Seiten in eine mehrseitige PDF zusammenführen
    if len(pages_data) == 1:
        buf = io.BytesIO()
        Image.open(io.BytesIO(pages_data[0])).save(buf, format="PDF", resolution=150.0)
        return buf.getvalue()

    first = Image.open(io.BytesIO(pages_data[0])).convert("RGB")
    rest  = [Image.open(io.BytesIO(p)).convert("RGB") for p in pages_data[1:]]
    out = io.BytesIO()
    first.save(out, format="PDF", resolution=150.0, save_all=True, append_images=rest)
    return out.getvalue()


def _apply_spreadsheet_print_settings(data: bytes, src_ext: str) -> bytes:
    """Entfernt Print_Area und setzt A4-Fit-to-Width direkt im xlsx-ZIP-XML.

    Zweistufig:
    1. ZIP/XML-Pass: _xlnm.Print_Area aus workbook.xml löschen + Row/Col-Breaks
       aus sheet XMLs entfernen. Direktes XML-Patching ist zuverlässiger als
       openpyxl-Propertysetter (kein Risiko durch fehlerhaftes Re-Serialisieren).
    2. openpyxl-Pass: fitToWidth=1, paperSize=A4, Orientation setzen.
       Nach dem XML-Pass liegt kein Print_Area mehr im File, sodass LibreOffice
       alle Spalten rendert.
    """
    import io, zipfile, re as _re

    # ── Phase 1: direktes ZIP/XML-Patching ───────────────────────────────────
    try:
        buf_in = io.BytesIO(data)
        buf_out = io.BytesIO()
        with zipfile.ZipFile(buf_in, 'r') as zin, \
             zipfile.ZipFile(buf_out, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename == 'xl/workbook.xml':
                    # _xlnm.Print_Area definedName vollständig entfernen
                    content = _re.sub(
                        rb'<definedName[^>]*name="[^"]*Print_Area[^"]*"[^>]*>.*?</definedName>\s*',
                        b'', content, flags=_re.DOTALL | _re.IGNORECASE,
                    )
                    # Auch self-closing Variante (seltener)
                    content = _re.sub(
                        rb'<definedName[^/]*Print_Area[^/]*/>\s*',
                        b'', content, flags=_re.IGNORECASE,
                    )
                elif _re.match(rb'xl/worksheets/sheet\d+\.xml$',
                               item.filename.encode()):
                    # Manuelle Zeilen-/Spalten-Breaks entfernen
                    content = _re.sub(
                        rb'<rowBreaks\b[^>]*/>', b'', content)
                    content = _re.sub(
                        rb'<rowBreaks\b[^>]*>.*?</rowBreaks>', b'',
                        content, flags=_re.DOTALL)
                    content = _re.sub(
                        rb'<colBreaks\b[^>]*/>', b'', content)
                    content = _re.sub(
                        rb'<colBreaks\b[^>]*>.*?</colBreaks>', b'',
                        content, flags=_re.DOTALL)
                zout.writestr(item, content)
        data = buf_out.getvalue()
        logger.debug("_apply_spreadsheet_print_settings: XML-Pass OK")
    except Exception as exc:
        logger.debug("XML-Phase fehlgeschlagen (harmlos): %s", exc)

    # ── Phase 2: openpyxl für fitToPage/orientation ──────────────────────────
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.page import PageSetupProperties

        wb = load_workbook(filename=io.BytesIO(data))
        for ws in wb.worksheets:
            ws.page_setup.paperSize  = ws.PAPERSIZE_A4
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            if ws.max_column and ws.max_row:
                ratio = ws.max_column / max(ws.max_row, 1)
                ws.page_setup.orientation = (
                    ws.ORIENTATION_LANDSCAPE if ratio > 1.5
                    else ws.ORIENTATION_PORTRAIT
                )
            if ws.sheet_properties.pageSetUpPr is None:
                ws.sheet_properties.pageSetUpPr = PageSetupProperties()
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        out = io.BytesIO()
        wb.save(out)
        logger.debug("_apply_spreadsheet_print_settings: openpyxl-Pass OK, %d sheets", len(wb.worksheets))
        return out.getvalue()
    except Exception as exc:
        logger.debug("openpyxl-Phase fehlgeschlagen: %s", exc)
        return data


def _pdf_page_width_pts(pdf_path: str) -> float | None:
    """Gibt die Breite der ersten Seite eines PDFs in Punkten zurück.

    Liest MediaBox direkt aus den rohen PDF-Bytes via Regex — kein
    externer Dependency. Gibt None zurück wenn nicht lesbar.
    """
    import re
    try:
        with open(pdf_path, "rb") as f:
            data = f.read(65536)  # ersten 64 KB reichen für MediaBox
        m = re.search(
            rb"/MediaBox\s*\[\s*([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s*\]",
            data,
        )
        if m:
            return float(m.group(3)) - float(m.group(1))
    except Exception:
        pass
    return None


def _convert_libreoffice(data: bytes, src_ext: str) -> bytes:
    """Office-Formate → PDF via `libreoffice --headless`.

    Spec:
      - Schreibt Eingabe in /tmp/<uuid>.<src_ext>
      - Ruft `libreoffice --headless --convert-to pdf --outdir /tmp/<uuid>_out <input>`
      - Liest die produzierte PDF zurück
      - Räumt danach /tmp auf
    """
    if not shutil.which("libreoffice") and not shutil.which("soffice"):
        raise ConversionError(
            "LibreOffice ist im Container nicht installiert — "
            "dieses Format kann nicht konvertiert werden."
        )
    binary = shutil.which("libreoffice") or shutil.which("soffice")

    # Spreadsheets (xlsx/xls/ods): Fit-to-1-page-wide + A4 via openpyxl setzen,
    # sonst druckt LibreOffice das Sheet auf endlos-breiten Seiten.
    if src_ext in ("xlsx", "xls", "ods"):
        data = _apply_spreadsheet_print_settings(data, src_ext)

    import uuid
    work = tempfile.mkdtemp(prefix="printix-conv-")
    try:
        in_path = os.path.join(work, f"input.{src_ext}")
        out_dir = os.path.join(work, "out")
        os.makedirs(out_dir, exist_ok=True)
        with open(in_path, "wb") as f:
            f.write(data)

        # LibreOffice braucht HOME (sonst Profil-Fehler in Containern)
        env = os.environ.copy()
        env["HOME"] = work
        proc = subprocess.run(
            [binary, "--headless", "--convert-to", "pdf",
             "--outdir", out_dir, in_path],
            # v6.7.40: 120s war beim ersten DOCX-Send zu knapp — LibreOffice-
            # Coldstart im Container frisst ~60-90s nur fürs Profil-Init, dann
            # kommt noch die eigentliche Konvertierung. 300s gibt genug Luft.
            env=env, timeout=300,
            capture_output=True,
        )
        if proc.returncode != 0:
            err = (proc.stderr or b"").decode("utf-8", errors="replace")[:500]
            raise ConversionError(f"LibreOffice-Konvertierung fehlgeschlagen: {err}")

        # PDF im out_dir finden
        produced = [f for f in os.listdir(out_dir) if f.lower().endswith(".pdf")]
        if not produced:
            raise ConversionError("LibreOffice produzierte keine PDF-Datei")
        out_path = os.path.join(out_dir, produced[0])

        # Spreadsheets: Ghostscript-Pass erzwingt A4-Ausgabe.
        # LibreOffice ignoriert fitToPage bei headless-Konvertierung und
        # rendert Sheets auf endlos-breiten Seiten.
        # WICHTIG: -dPDFFitPage nur wenn Seite BREITER als A4 — sonst
        # werden schmale Sheets (z.B. 3 Spalten) auf A4-Breite hochskaliert
        # und Zellen erscheinen riesig im Preview.
        if src_ext in ("xlsx", "xls", "ods") and shutil.which("gs"):
            gs_out = os.path.join(out_dir, "a4.pdf")
            # Seitenbreite des LibreOffice-PDFs ermitteln (in Punkten)
            pdf_width_pts = _pdf_page_width_pts(out_path)
            A4_W_PTS = 595.28  # A4 Breite in Punkten (72dpi)
            needs_scale_down = pdf_width_pts is None or pdf_width_pts > A4_W_PTS * 1.05
            gs_args = ["gs", "-dBATCH", "-dNOPAUSE", "-dSAFER", "-dQUIET",
                       "-sDEVICE=pdfwrite", "-dFIXEDMEDIA", "-sPAPERSIZE=a4"]
            if needs_scale_down:
                gs_args.append("-dPDFFitPage")  # nur verkleinern wenn nötig
            gs_args.append(f"-sOutputFile={gs_out}")
            gs_args.append(out_path)
            gs_proc = subprocess.run(gs_args, timeout=60, capture_output=True)
            if gs_proc.returncode == 0 and os.path.exists(gs_out):
                out_path = gs_out
            else:
                logger.warning(
                    "_convert_libreoffice: gs A4-Pass fehlgeschlagen (%d), "
                    "Original-PDF wird verwendet",
                    gs_proc.returncode,
                )

        with open(out_path, "rb") as f:
            return f.read()
    finally:
        try:
            shutil.rmtree(work, ignore_errors=True)
        except Exception:
            pass


# ─── Orchestrierung ──────────────────────────────────────────────────────────

def convert_to_pdf(data: bytes, filename: str = "",
                   bw: bool = False, image_size: str = "full") -> tuple[bytes, str]:
    """Haupt-Entry: erkennt Format, ruft passenden Konverter auf.

    bw=True        Bild in Graustufen konvertieren (nur für image/*).
    image_size     Druckgröße: "full" | "10x13" | "13x18" | "original".

    Returns: (pdf_bytes, source_format_label)
    Raises: ConversionError falls das Format nicht konvertierbar ist
    """
    mime, ext = detect_format(data, filename)
    logger.info("Upload-Konverter: Input erkannt — mime=%s ext=%s size=%d bw=%s img_size=%s",
                mime, ext, len(data), bw, image_size)

    if mime == "application/pdf":
        return data, "pdf (passthrough)"

    if mime.startswith("image/"):
        pdf = _convert_image_to_pdf(data, bw=bw, image_size=image_size)
        return pdf, f"image ({ext}) → pdf"

    if mime == "text/plain":
        pdf = _convert_text_to_pdf(data)
        return pdf, "text → pdf"

    # Office-Formate + RTF → LibreOffice
    libreoffice_formats = {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "application/vnd.oasis.opendocument.text",
        "application/vnd.oasis.opendocument.spreadsheet",
        "application/vnd.oasis.opendocument.presentation",
        "application/msword",
        "application/vnd.ms-excel",
        "application/vnd.ms-powerpoint",
        "application/rtf",
    }
    if mime in libreoffice_formats or ext in ("docx","xlsx","pptx","odt","ods","odp","doc","xls","ppt","rtf"):
        pdf = _convert_libreoffice(data, ext or "bin")
        return pdf, f"libreoffice ({ext or mime}) → pdf"

    raise ConversionError(
        f"Format nicht unterstützt: mime={mime} ext={ext}. "
        f"Erlaubt: PDF, docx/xlsx/pptx, odt/ods/odp, rtf, TXT, png/jpg/gif/bmp/tiff."
    )


def is_libreoffice_available() -> bool:
    """Für UI-Hinweise nützlich."""
    return bool(shutil.which("libreoffice") or shutil.which("soffice"))


def render_image_preview_png(img_bytes: bytes, max_size: int = 900) -> Optional[bytes]:
    """Rendert ein Bild (JPEG/PNG/GIF/BMP/TIFF) als PNG-Thumbnail via Pillow.

    Gibt None zurück wenn die Daten kein unterstütztes Bildformat sind oder
    Pillow nicht verfügbar ist.
    """
    if not img_bytes or img_bytes[:4] == b"%PDF":
        return None
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")
        img.thumbnail((max_size, max_size), Image.LANCZOS)
        out = io.BytesIO()
        img.save(out, format="PNG")
        return out.getvalue()
    except Exception as exc:
        logger.debug("render_image_preview_png failed: %s", exc)
    return None


def render_preview_png(pdf_bytes: bytes, dpi: int = 96) -> Optional[bytes]:
    """Rendert Seite 1 eines PDFs als PNG via Ghostscript.

    Gibt None zurück wenn Ghostscript nicht verfügbar, das PDF leer/kaputt
    ist oder das Rendering länger als 30s dauert.
    """
    if not pdf_bytes or pdf_bytes[:4] != b"%PDF":
        return None
    gs = shutil.which("gs")
    if not gs:
        return None
    try:
        with tempfile.TemporaryDirectory(prefix="msp-prev-") as tmp:
            in_path = os.path.join(tmp, "in.pdf")
            out_path = os.path.join(tmp, "page001.png")
            with open(in_path, "wb") as fh:
                fh.write(pdf_bytes)
            subprocess.run(
                [gs, "-dBATCH", "-dNOPAUSE", "-dSAFER", "-dQUIET",
                 "-sDEVICE=png16m", f"-r{dpi}",
                 "-dFirstPage=1", "-dLastPage=1",
                 f"-sOutputFile={out_path}", in_path],
                timeout=30, check=True, capture_output=True,
            )
            if os.path.exists(out_path):
                with open(out_path, "rb") as fh:
                    return fh.read()
    except Exception as exc:
        logger.debug("render_preview_png failed: %s", exc)
    return None
